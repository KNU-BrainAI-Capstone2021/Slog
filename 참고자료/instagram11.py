# 인스타그램 셀레니움으로 크롤링`
# 4. 뽑은 아이디로 각자의 프로필화면으로 들어가 프로필 다운받기.

import requests
import urllib.request
import random
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time

import sys
non_bmp_map = dict.fromkeys(range(0x10000, sys.maxunicode + 1), 0xfffd)



# 엑셀 파일로 저장된 id(B1~B200) 있으면 그걸 arr_id 에 넣는다.
arr_profile = []                                                                        # 프로필 화면 주소들 담김.
arr_id = []                                                                             # 아이디 담김.

# 엑셀 파일 불러옴, id로 프로필 주소 만들기
# https://www.instagram.com/(id 넣음)/?hl=ko 형식
from openpyxl import load_workbook

wb = load_workbook("instagramWithForm.xlsx")
ws = wb.active
for i in range(2,202):
    id = ws["B"+str(i)].value
    profile = "https://www.instagram.com/"+ id +"/?hl=ko"
    print("[%d] : %s"%(i,id))
    arr_id.append(id)
    arr_profile.append(profile)
    
    
print("=================================================================")
print("id 수 : %d, 프로필 주소 수 : %d" %(len(arr_id), len(arr_profile)))


# 드라이버 연결
driver = webdriver.Chrome(executable_path="D:\\python_D\\chromedriver.exe")

count = 0
count_img = 0
# 4. 뽑은 아이디로 각자의 프로필화면으로 들어가 프로필 다운받기.
for profile in arr_profile:
    if(arr_id[count] != "none"):
        driver.implicitly_wait(2)
        driver.get(profile)     # 프로필화면으로 들어감

        # 프로필 들어가서 이미지 검색
        try:   
            href = driver.find_elements_by_xpath('//*[@class="_2dbep "]')                       # <div class="_2dbep"> <img src="주소"> 형식.
            print("프로필 클래스 검색성공", len(href))
            print(" profile[%d]" % count)
        except:
            print("프로필 클래스 검색실패")

        # 이미지 저장.
        try:
            full_name = "D:\\python_D\\크롤러\\오늘의훈남\\" + arr_id[count] + ".jpg"
            print("저장될 이미지 : ",full_name)
            href_profile = href[0].find_element_by_css_selector('img').get_attribute('src')      # 프로필 사진 검색
            print("href_profile : ", href_profile)
#            urllib.request.urlretrieve(item.get_attribute('src'), full_name) # src를 받는다.
            urllib.request.urlretrieve(href_profile, full_name)                                  # full_name에 받아온 이미지 저장.
        
            print("img[%d] : %s" %(count, arr_id[count]+".jpg"))
        except:
            print("이미지 저장 실패")
            
        count_img += 1    
        print("-------------------------")
        
    count += 1
    
driver.quit()
print("이미지 개수 : %d"%count_img)
print("Saved!")
